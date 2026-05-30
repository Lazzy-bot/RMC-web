import os
import sys
import openpyxl

# Add backend to sys.path
sys.path.append(os.path.join(os.getcwd(), "backend"))

from services.excel import _resolve_excel_item, _download_excel
from config import CHART_EXCEL_SHARE_LINK

def debug_read():
    print(f"Resolving: {CHART_EXCEL_SHARE_LINK}")
    try:
        # We need to be logged in for this to work in a script
        # But maybe the token is cached in BASE_DIR/Cache/token_cache.bin
        drive_id, item_id = _resolve_excel_item(CHART_EXCEL_SHARE_LINK)
        print(f"Drive ID: {drive_id}")
        print(f"Item ID: {item_id}")
        
        path = _download_excel(drive_id, item_id)
        print(f"Downloaded to: {path}")
        
        wb = openpyxl.load_workbook(path, data_only=True)
        print(f"Sheets: {wb.sheetnames}")
        
        ws = wb.active
        print(f"Active Sheet: {ws.title}")
        
        print("Rows 1-10 (Col A):")
        for r in range(1, 11):
            val = ws.cell(row=r, column=1).value
            print(f"Row {r}: {val}")
            
        os.unlink(path)
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    debug_read()
